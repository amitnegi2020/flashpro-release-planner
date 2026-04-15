import sys, json, zipfile, os, tempfile

rp_path = sys.argv[1]

try:
    with zipfile.ZipFile(rp_path, 'r') as zf:
        names = zf.namelist()

        # Best case: state.json contains everything
        if 'state.json' in names:
            with zf.open('state.json') as f:
                state = json.loads(f.read())
            # Also load streams from streams.json if present (more reliable)
            if 'streams.json' in names:
                with zf.open('streams.json') as f:
                    state['streams'] = json.loads(f.read())
            # Load manifest for project name
            project_name = state.get('projectName', 'Imported Project')
            if 'manifest.json' in names:
                with zf.open('manifest.json') as f:
                    manifest = json.loads(f.read())
                project_name = manifest.get('projectName', project_name)
            print(json.dumps({'ok': True, 'projectName': project_name, 'state': state}))
        else:
            # Fallback: reconstruct from individual files
            state = {}
            if 'release_plan.xlsx' in names:
                # Extract and parse xlsx
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    tmp.write(zf.read('release_plan.xlsx'))
                    tmp_path = tmp.name
                import openpyxl, re
                wb = openpyxl.load_workbook(tmp_path)
                os.unlink(tmp_path)
                ws = wb['Release Plan'] if 'Release Plan' in wb.sheetnames else wb[wb.sheetnames[0]]
                headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
                def find(names_):
                    for n in names_:
                        for i,h in enumerate(headers):
                            if h and n.lower() in h.lower(): return i
                    return None
                sid_idx = find(['Story ID'])
                hl_idx = find(['Headline'])
                persona_idx = find(['Persona'])
                goal_idx = find(['Goal'])
                action_idx = find(['Action'])
                pts_idx = find(['Points'])
                sprint_idx = find(['Sprint'])
                stream_idx = find(['Stream Name','Stream'])
                stories = []
                seen = set()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    sid = str(row[sid_idx]).strip() if sid_idx is not None and row[sid_idx] else None
                    if not sid or sid in seen or sid == 'None': continue
                    seen.add(sid)
                    sprint_raw = str(row[sprint_idx]).strip() if sprint_idx is not None and row[sprint_idx] else 'Backlog'
                    sprint_val = 'backlog' if sprint_raw.lower() == 'backlog' else 'sprint' + re.search(r'(\d+)', sprint_raw).group(1) if re.search(r'(\d+)', sprint_raw) else 'backlog'
                    stories.append({
                        'id': sid,
                        'headline': str(row[hl_idx]).strip() if hl_idx is not None and row[hl_idx] else '',
                        'persona': str(row[persona_idx]).strip() if persona_idx is not None and row[persona_idx] else '',
                        'goal': str(row[goal_idx]).strip() if goal_idx is not None and row[goal_idx] else '',
                        'action': str(row[action_idx]).strip() if action_idx is not None and row[action_idx] else '',
                        'pts': float(row[pts_idx]) if pts_idx is not None and row[pts_idx] else 0,
                        'sprint': sprint_val,
                        'stream': 'backlog',
                    })
                state['stories'] = stories
            if 'streams.json' in names:
                with zf.open('streams.json') as f:
                    state['streams'] = json.loads(f.read())
            if 'dependencies.json' in names:
                with zf.open('dependencies.json') as f:
                    state['customDeps'] = json.loads(f.read())
            if 'critical_path.json' in names:
                with zf.open('critical_path.json') as f:
                    state['criticalPath'] = json.loads(f.read())
            project_name = 'Imported Project'
            if 'manifest.json' in names:
                with zf.open('manifest.json') as f:
                    manifest = json.loads(f.read())
                project_name = manifest.get('projectName', project_name)
            state['projectName'] = project_name
            # BUG-10 fix: validate that we actually reconstructed usable state
            if not state.get('stories') and 'release_plan.xlsx' not in names and 'state.json' not in names:
                print(json.dumps({'ok': False, 'error': 'This does not appear to be a valid FlashPro project file (.rp). Required files not found in archive.'}))
            else:
                print(json.dumps({'ok': True, 'projectName': project_name, 'state': state}))
except Exception as e:
    import traceback
    print(json.dumps({'ok': False, 'error': str(e), 'trace': traceback.format_exc()}))
