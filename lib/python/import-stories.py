import openpyxl, json, sys, re

xlsx_path = sys.argv[1]

try:
    wb = openpyxl.load_workbook(xlsx_path)
    sheet_names = list(wb.sheetnames)

    # Feature 9: Detect if this is a Release Plan export
    is_export = 'Release Plan' in sheet_names and 'Board - Stories' in sheet_names

    if is_export:
        ws = wb['Release Plan']
    else:
        ws = wb[sheet_names[0]]

    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

    # ── Dynamic column resolution ─────────────────────────────────────────────
    # Instead of hardcoding field names, we find key columns by semantic meaning
    # and then store ALL other columns dynamically on each story.
    # This means Workflow, Platform Capability, or any future column is auto-captured.

    def find_col(candidates):
        "Find first matching column index by partial name match (case-insensitive)"
        for name in candidates:
            for i, h in enumerate(headers):
                if h and name.lower() in h.lower(): return i
        return None

    def find_exact(name):
        for i, h in enumerate(headers):
            if h and h.lower() == name.lower(): return i
        return None

    def to_key(col_name):
        "Convert column header to a safe camelCase key for the story object"
        import re as _re
        words = _re.split(r'[\s_\-/()]+', col_name.strip())
        if not words: return col_name.lower()
        return words[0].lower() + ''.join(w.capitalize() for w in words[1:] if w)

    # Identify the reserved system columns by role
    sid_idx       = find_col(['Story ID', 'id'])
    headline_idx  = find_col(['Headline', 'title', 'name', 'summary'])
    status_idx    = find_col(['Status'])
    pts_col       = find_col(['Story Estimate (weeks)', 'Story Estimate', 'Points', 'pts', 'estimate'])
    original_idx  = find_col(['Original Story ID', 'Parent Story', 'Derived From', 'Split From'])
    sprint_col    = find_exact('Sprint')
    stream_col    = find_exact('Stream')

    # BUG-11 fix: explicit error if Story ID column is absent
    if sid_idx is None:
        print(json.dumps({'error': 'File must have a Story ID column. Check this is a valid Story Map file.'}))
        sys.exit(0)

    # Build a map of ALL other columns → story key (dynamic, no hardcoding)
    SYSTEM_COLS = {'id', 'story id', 'sprint', 'stream', 'stream name', 'stream number'}
    extra_cols = []
    for i, h in enumerate(headers):
        if not h: continue
        if h.lower() in SYSTEM_COLS: continue
        if i in [sid_idx, sprint_col, stream_col]: continue
        extra_cols.append((i, h, to_key(h)))

    # First pass: collect all split story IDs
    split_story_ids = set()
    if status_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[status_idx]:
                st = str(row[status_idx]).strip().lower()
                sid = str(row[sid_idx]).strip() if sid_idx is not None and row[sid_idx] else None
                if st == 'split' and sid and sid != 'None':
                    split_story_ids.add(sid)

    stories = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[sid_idx]).strip() if sid_idx is not None and row[sid_idx] else None
        if not sid or sid in seen or sid == 'None': continue
        seen.add(sid)

        # Skip deleted / split parent stories
        if status_idx is not None and row[status_idx]:
            st = str(row[status_idx]).strip().lower()
            if st in ('delete', 'deleted', 'split'): continue

        sprint_val = 'backlog'
        stream_val = 'backlog'

        original_story_id = None
        if original_idx is not None and row[original_idx]:
            original_story_id = str(row[original_idx]).strip()
            if original_story_id == 'None': original_story_id = None

        if is_export and sprint_col is not None and row[sprint_col]:
            raw = str(row[sprint_col]).strip()
            if raw.lower() == 'backlog':
                sprint_val = 'backlog'
            else:
                m = re.search(r'(\d+)', raw)
                if m: sprint_val = 'sprint' + m.group(1)

        if is_export and stream_col is not None and row[stream_col]:
            raw = str(row[stream_col]).strip()
            if raw.lower() in ('unassigned', 'backlog'):
                stream_val = 'backlog'
            else:
                m = re.search(r'(\d+)', raw)
                if m: stream_val = 's' + m.group(1)

        pts = None
        if pts_col is not None and row[pts_col] is not None:
            try: pts = float(row[pts_col])
            except: pts = None

        # Build story with ALL columns dynamically — no hardcoding
        story = {
            'id': sid,
            'sprint': sprint_val,
            'stream': stream_val,
            'pts': pts,
            'originalStoryId': original_story_id,
            'isSplitChild': original_story_id is not None and original_story_id in split_story_ids,
        }

        # Add every column from the sheet dynamically
        for (col_i, col_name, col_key) in extra_cols:
            if col_key in ('pts', 'id', 'sprint', 'stream'): continue
            val = row[col_i]
            if val is not None and str(val).strip() not in ('', 'None'):
                story[col_key] = str(val).strip()
            else:
                story[col_key] = ''

        # Ensure backward-compatible aliases for common fields
        story['headline'] = story.get('headline') or story.get('title') or story.get('name') or story.get('summary') or ''
        story['persona']  = story.get('persona') or story.get('userPersona') or ''
        story['goal']     = story.get('goal') or ''
        story['action']   = story.get('action') or ''
        story['capability'] = story.get('capability') or story.get('platformCapability') or ''
        story['workflow'] = story.get('workflow') or ''
        story['status']   = story.get('status') or ''

        stories.append(story)

    # Return ALL column headers as available_columns (including dynamic ones)
    available_columns = [h for (_, h, _) in extra_cols if h]

    print(json.dumps({'stories': stories, 'is_export': is_export, 'columns': available_columns, 'sheet_names': sheet_names}))
except Exception as e:
    import traceback
    print(json.dumps({'error': str(e), 'trace': traceback.format_exc()}))
