import openpyxl, json, sys
try:
    wb = openpyxl.load_workbook(sys.argv[1])
    # Try "Dependencies" sheet first
    ws = wb['Dependencies'] if 'Dependencies' in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    sid_col = next((i for i,h in enumerate(headers) if 'story id' in h.lower() and 'depends' not in h.lower()), None)
    dep_col = next((i for i,h in enumerate(headers) if 'depends' in h.lower()), None)
    if sid_col is None or dep_col is None:
        print(json.dumps({'error': 'Could not find Story ID and Depends On columns'}))
        exit()
    deps = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[sid_col]).strip() if row[sid_col] else None
        dep = str(row[dep_col]).strip() if row[dep_col] else None
        if not sid or sid == 'None' or not dep or dep == 'None': continue
        if sid not in deps: deps[sid] = []
        if dep not in deps[sid]: deps[sid].append(dep)
    # Also try "Critical Path" sheet for CP list
    cp = []
    if 'Critical Path' in wb.sheetnames:
        ws2 = wb['Critical Path']
        h2 = [str(c.value).strip() if c.value else '' for c in ws2[1]]
        sid2 = next((i for i,h in enumerate(h2) if 'story id' in h.lower()), 1)
        for row in ws2.iter_rows(min_row=2, values_only=True):
            v = str(row[sid2]).strip() if row[sid2] else None
            if v and v != 'None' and not v.startswith('Phase'): cp.append(v)
    print(json.dumps({'deps': deps, 'cp': cp}))
except Exception as e:
    import traceback
    print(json.dumps({'error': str(e), 'trace': traceback.format_exc()}))
