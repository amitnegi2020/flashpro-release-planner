import json, sys, zipfile, os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

data_path = sys.argv[1]
out_path = sys.argv[2]

with open(data_path) as f:
    data = json.load(f)

stories = data['stories']
streams = data['streams']
custom_deps = data.get('customDeps', {})
critical_path = data.get('criticalPath', [])
project_name = data.get('projectName', 'Release Plan')
saved_at = data.get('savedAt', datetime.utcnow().isoformat())

stream_map = {s['key']: s for s in streams}
stream_num = {s['key']: i+1 for i, s in enumerate(streams)}

SPRINTS = [
    {'key': 'sprint1', 'label': 'Sprint 1'}, {'key': 'sprint2', 'label': 'Sprint 2'},
    {'key': 'sprint3', 'label': 'Sprint 3'}, {'key': 'sprint4', 'label': 'Sprint 4'},
    {'key': 'sprint5', 'label': 'Sprint 5'}, {'key': 'sprint6', 'label': 'Sprint 6'},
    {'key': 'sprint7', 'label': 'Sprint 7'}, {'key': 'sprint8', 'label': 'Sprint 8'},
]

tmp_dir = os.path.dirname(out_path)

# ── 1. Release Plan xlsx ────────────────────────────────────────────────────
def lighten(hex_color, factor=0.85):
    h = hex_color.lstrip('#')
    r,g,b = int(h[0:2],16),int(h[2:4],16),int(h[4:6],16)
    return 'FF%02X%02X%02X' % (int(r+(255-r)*factor), int(g+(255-g)*factor), int(b+(255-b)*factor))

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = 'Release Plan'
hdr = ['Sprint','Stream Name','Stream Number','Story ID','Headline','Persona','Goal','Action','Points']
ws1.append(hdr)
for c in ws1[1]:
    c.font = Font(bold=True, color='FFFFFFFF')
    c.fill = PatternFill('solid', fgColor='FF0F172A')

sprintOrder = ['backlog','sprint1','sprint2','sprint3','sprint4','sprint5','sprint6','sprint7','sprint8']
sorted_stories = sorted(stories, key=lambda s: (sprintOrder.index(s.get('sprint','backlog')) if s.get('sprint','backlog') in sprintOrder else 99, s.get('stream',''), s.get('id','')))

for s in sorted_stories:
    sn = stream_map.get(s.get('stream',''), {})
    snum = stream_num.get(s.get('stream',''), '')
    sprint_raw = s.get('sprint','')
    sprint_label = 'Backlog' if sprint_raw == 'backlog' else sprint_raw.replace('sprint','Sprint ')
    stream_name = 'Unassigned' if s.get('stream','') in ('backlog','') else ('Stream '+str(snum)+' — '+sn.get('name','') if sn and snum else 'Stream '+str(snum))
    stream_num_label = 'Stream '+str(snum) if snum else 'Unassigned'
    ws1.append([sprint_label, stream_name, stream_num_label, s.get('id',''), s.get('headline',''), s.get('persona',''), s.get('goal',''), s.get('action',''), s.get('pts','')])

for col_idx, col in enumerate(ws1.columns, 1):
    ml = max((len(str(c.value or '')) for c in col), default=10)
    ws1.column_dimensions[get_column_letter(col_idx)].width = min(ml+2, 50)

# Board sheets
def make_board(wb, title, cell_fn):
    ws = wb.create_sheet(title)
    ws.append(['Stream / Sprint'] + [sp['label'] for sp in SPRINTS])
    for ci, c in enumerate(ws[1], 1):
        c.font = Font(bold=True, color='FFFFFFFF')
        c.fill = PatternFill('solid', fgColor='FF334155' if ci==1 else 'FF1E3A5F')
        c.alignment = Alignment(horizontal='center')
    for si in streams:
        snum_b = stream_num.get(si['key'],'')
        row = [('Stream '+str(snum_b)+' — '+si['name']) if snum_b else si['name']] + [cell_fn(sp['key'], si['key'], stories) for sp in SPRINTS]
        ws.append(row)
        r = ws.max_row
        lf = PatternFill('solid', fgColor=lighten(si.get('color','#64748b')))
        ws.cell(r,1).font = Font(bold=True)
        ws.cell(r,1).fill = lf
        for ci in range(2, len(SPRINTS)+2):
            ws.cell(r,ci).alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['A'].width = 22
    for ci in range(2, len(SPRINTS)+2):
        ws.column_dimensions[get_column_letter(ci)].width = 28

def stories_cell(sk, stk, ss):
    items=[s for s in ss if s.get('sprint')==sk and s.get('stream')==stk]
    return '\n'.join(f"{s['id']}: {s.get('headline','')[:40]}" for s in items) if items else ''

def actions_cell(sk, stk, ss):
    items=[s for s in ss if s.get('sprint')==sk and s.get('stream')==stk]
    if not items: return ''
    g={}
    for s in items:
        k=f"{s.get('action','')} ({s.get('persona','')})"
        if k not in g: g[k]={'c':0,'p':0}
        g[k]['c']+=1; g[k]['p']+=float(s.get('pts') or 0)
    return '\n'.join(f"{k}: {v['c']}s, {v['p']:.1f}pts" for k,v in g.items())

def goals_cell(sk, stk, ss):
    items=[s for s in ss if s.get('sprint')==sk and s.get('stream')==stk]
    if not items: return ''
    g={}
    for s in items:
        k=s.get('goal','')
        if k not in g: g[k]={'c':0,'p':0}
        g[k]['c']+=1; g[k]['p']+=float(s.get('pts') or 0)
    return '\n'.join(f"{k}: {v['c']}s, {v['p']:.1f}pts" for k,v in g.items())

make_board(wb, 'Board - Stories', stories_cell)
make_board(wb, 'Board - Actions', actions_cell)
make_board(wb, 'Board - Goals', goals_cell)

ws5 = wb.create_sheet('Stream Config')
ws5.append(['Stream Number','Stream Name','Stream Description','Stream Color'])
for c in ws5[1]:
    c.font = Font(bold=True, color='FFFFFFFF')
    c.fill = PatternFill('solid', fgColor='FF0F172A')
for i, s in enumerate(streams, 1):
    ws5.append([i, s.get('name',''), s.get('description',''), s.get('color','')])

xlsx_path = os.path.join(tmp_dir, 'release_plan.xlsx')
wb.save(xlsx_path)

# ── 2. Dependencies JSON ─────────────────────────────────────────────────────
deps_path = os.path.join(tmp_dir, 'dependencies.json')
with open(deps_path, 'w') as f:
    json.dump(custom_deps, f, indent=2)

# ── 3. Critical Path JSON ────────────────────────────────────────────────────
cp_path = os.path.join(tmp_dir, 'critical_path.json')
with open(cp_path, 'w') as f:
    json.dump(critical_path, f, indent=2)

# ── 4. Streams config JSON ───────────────────────────────────────────────────
streams_path = os.path.join(tmp_dir, 'streams.json')
with open(streams_path, 'w') as f:
    json.dump(streams, f, indent=2)

# ── 5. Project manifest JSON ─────────────────────────────────────────────────
manifest = {
    'projectName': project_name,
    'savedAt': saved_at,
    'exportedAt': datetime.utcnow().isoformat() + 'Z',
    'format': 'MakroPro Release Planner .RP v1',
    'storyCount': len(stories),
    'plannedCount': len([s for s in stories if s.get('sprint','backlog') != 'backlog']),
    'streams': [{'key': s['key'], 'name': s['name']} for s in streams],
    'files': {
        'release_plan': 'release_plan.xlsx',
        'dependencies': 'dependencies.json',
        'critical_path': 'critical_path.json',
        'streams': 'streams.json',
        'state': 'state.json'
    },
    'instructions': [
        '1. In MakroPro Release Planner, click Import > Import Release Plan',
        '2. Select release_plan.xlsx — restores full board with sprint + stream assignments',
        '3. Click Deps button (upload) > select dependencies.json — restores dependency rules',
        '4. Streams are restored automatically from streams.json embedded in the plan',
        '5. Project name and configuration are preserved',
    ]
}
manifest_path = os.path.join(tmp_dir, 'manifest.json')
with open(manifest_path, 'w') as f:
    json.dump(manifest, f, indent=2)

# ── 6. Full state JSON (complete restore in one step) ───────────────────────
state_path = os.path.join(tmp_dir, 'state.json')
with open(state_path, 'w') as f:
    json.dump(data, f, indent=2)

# ── Bundle into .RP zip ──────────────────────────────────────────────────────
with zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as zf:
    zf.write(xlsx_path, 'release_plan.xlsx')
    zf.write(deps_path, 'dependencies.json')
    zf.write(cp_path, 'critical_path.json')
    zf.write(streams_path, 'streams.json')
    zf.write(manifest_path, 'manifest.json')
    zf.write(state_path, 'state.json')

# Cleanup temp files
for p in [xlsx_path, deps_path, cp_path, streams_path, manifest_path, state_path]:
    try: os.unlink(p)
    except: pass

print('ok')
