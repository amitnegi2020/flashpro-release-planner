import openpyxl, json, sys, re

xlsx_path = sys.argv[1]
existing_json = sys.argv[2]
required_columns_json = sys.argv[3] if len(sys.argv) > 3 else '[]'

try:
    existing = json.loads(existing_json)
    existing_map = {s['id']: s for s in existing}

    # F8G.01/F8G.02 fix: required columns from the original story map
    required_columns = json.loads(required_columns_json)  # list of column header strings

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    headers_lower = [h.lower() for h in headers]

    def find_col(candidates):
        for name in candidates:
            for i, h in enumerate(headers):
                if h and name.lower() in h.lower(): return i
        return None

    sid_idx = find_col(['Story ID', 'id'])
    hl_idx  = find_col(['Headline', 'title', 'name', 'summary'])
    st_idx  = find_col(['Status'])

    if sid_idx is None:
        print(json.dumps({'error': 'File must have a Story ID column. Check this is a valid Story Map file.'}))
        exit()

    # F8G.01/F8G.02: Check that required columns from original story map are present
    # Skip standard system columns — only check dynamic aggregate columns
    SYSTEM_COLS_LOWER = {'story id', 'id', 'headline', 'title', 'name', 'summary',
                          'user persona', 'persona', 'goal', 'action', 'status',
                          'story estimate (weeks)', 'story estimate', 'points', 'pts', 'estimate',
                          'platform capability', 'capability', 'workflow', 'original story id',
                          'sprint', 'stream'}
    missing_cols = []
    for req in required_columns:
        if req.lower() in SYSTEM_COLS_LOWER:
            continue  # system columns are optional in AB files
        # Check if incoming file has a partial match for this column
        found = any(req.lower() in h for h in headers_lower)
        if not found:
            missing_cols.append(req)

    if missing_cols:
        cols_str = ', '.join(missing_cols)
        print(json.dumps({'error': f'Required columns missing from file: {cols_str}. Add these columns and re-import.', 'missingColumns': missing_cols}))
        exit()

    conflicts = []
    new_count = 0
    seen = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[sid_idx]).strip() if row[sid_idx] else None
        if not sid or sid in seen or sid == 'None': continue
        # Skip deleted/split
        if st_idx is not None and row[st_idx]:
            st = str(row[st_idx]).strip().lower()
            if st in ('delete', 'deleted', 'split'): continue
        seen.add(sid)

        incoming_hl = str(row[hl_idx]).strip() if hl_idx is not None and row[hl_idx] else ''

        if sid in existing_map:
            ex = existing_map[sid]
            ex_hl = ex.get('headline', '') or ex.get('title', '') or ''
            ex_sprint = ex.get('sprint', 'backlog')

            issues = []
            if ex_hl and incoming_hl and ex_hl.lower() != incoming_hl.lower():
                issues.append('headline_mismatch')
            if ex_sprint not in ('backlog', '', None):
                issues.append('already_planned')

            if issues:
                conflicts.append({
                    'id': sid,
                    'issues': issues,
                    'existingHeadline': ex_hl,
                    'incomingHeadline': incoming_hl,
                    'existingSprint': ex_sprint,
                })
        else:
            new_count += 1

    print(json.dumps({'ok': True, 'conflicts': conflicts, 'newCount': new_count}))
except Exception as e:
    import traceback
    print(json.dumps({'error': str(e), 'trace': traceback.format_exc()}))
