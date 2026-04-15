import json, openpyxl, sys
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

data_path = sys.argv[1]
out_path = sys.argv[2]

with open(data_path) as f:
    data = json.load(f)

stories = data['stories']
streams = data['streams']

stream_map = {s['key']: s for s in streams}
stream_num = {s['key']: i+1 for i, s in enumerate(streams)}

SPRINTS = [
    {'key': 'sprint1', 'label': 'Sprint 1', 'color': '1D4ED8'},
    {'key': 'sprint2', 'label': 'Sprint 2', 'color': '4338CA'},
    {'key': 'sprint3', 'label': 'Sprint 3', 'color': '6D28D9'},
    {'key': 'sprint4', 'label': 'Sprint 4', 'color': '7C3AED'},
    {'key': 'sprint5', 'label': 'Sprint 5', 'color': 'DB2777'},
    {'key': 'sprint6', 'label': 'Sprint 6', 'color': 'D97706'},
    {'key': 'sprint7', 'label': 'Sprint 7', 'color': '059669'},
    {'key': 'sprint8', 'label': 'Sprint 8', 'color': '0891B2'},
]

def hex_to_argb(hex_color):
    h = hex_color.lstrip('#')
    return 'FF' + h.upper()

def lighten_hex(hex_color, factor=0.85):
    h = hex_color.lstrip('#')
    r, g, b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
    r = int(r + (255-r)*factor)
    g = int(g + (255-g)*factor)
    b = int(b + (255-b)*factor)
    return 'FF%02X%02X%02X' % (r, g, b)

wb = openpyxl.Workbook()

# Sheet 1: Release Plan
ws1 = wb.active
ws1.title = 'Release Plan'
headers = ['Sprint', 'Stream Name', 'Stream Number', 'Story ID', 'Headline', 'Persona', 'Goal', 'Action', 'Points']
ws1.append(headers)
for cell in ws1[1]:
    cell.font = Font(bold=True, color='FFFFFFFF')
    cell.fill = PatternFill('solid', fgColor='FF1E3A5F')

for s in stories:
    sn = stream_map.get(s.get('stream',''), {})
    snum = stream_num.get(s.get('stream',''), '')
    sprint_raw = s.get('sprint','')
    sprint_label = 'Backlog' if sprint_raw == 'backlog' else sprint_raw.replace('sprint', 'Sprint ')
    stream_label = 'Unassigned' if s.get('stream','') in ('backlog','') else (sn.get('name','') if sn else ('Stream ' + str(snum)))
    ws1.append([
        sprint_label,
        ('Stream ' + str(snum) + ' — ' + sn.get('name','')) if sn and snum else stream_label,
        'Stream ' + str(snum) if snum else 'Unassigned',
        s.get('id',''),
        s.get('headline',''),
        s.get('persona',''),
        s.get('goal',''),
        s.get('action',''),
        s.get('pts','')
    ])

for col_idx, col in enumerate(ws1.columns, 1):
    max_len = max((len(str(c.value or '')) for c in col), default=10)
    ws1.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

def make_board_sheet(wb, title, cell_fn):
    ws = wb.create_sheet(title)
    header_row = ['Stream / Sprint'] + [sp['label'] for sp in SPRINTS]
    ws.append(header_row)
    for ci, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color='FFFFFFFF')
        if ci == 1:
            cell.fill = PatternFill('solid', fgColor='FF334155')
        else:
            sp = SPRINTS[ci-2]
            cell.fill = PatternFill('solid', fgColor=hex_to_argb(sp['color']))
        cell.alignment = Alignment(horizontal='center')

    for s_info in streams:
        sk = s_info['key']
        color = s_info.get('color', '#64748b')
        snum_b = stream_num.get(s_info['key'],'')
        row_data = [('Stream ' + str(snum_b) + ' — ' + s_info['name']) if snum_b else s_info['name']]
        for sp in SPRINTS:
            content = cell_fn(sp['key'], sk, stories)
            row_data.append(content)
        ws.append(row_data)
        row_num = ws.max_row
        stream_fill = PatternFill('solid', fgColor=lighten_hex(color))
        name_cell = ws.cell(row=row_num, column=1)
        name_cell.font = Font(bold=True)
        name_cell.fill = stream_fill
        for ci in range(2, len(SPRINTS)+2):
            c = ws.cell(row=row_num, column=ci)
            c.alignment = Alignment(wrap_text=True, vertical='top')

    ws.column_dimensions['A'].width = 22
    for ci in range(2, len(SPRINTS)+2):
        ws.column_dimensions[get_column_letter(ci)].width = 28
    return ws

def stories_cell(sprint_key, stream_key, stories):
    items = [s for s in stories if s.get('sprint') == sprint_key and s.get('stream') == stream_key]
    if not items: return ''
    return '\n'.join(f"{s['id']}: {s.get('headline','')[:40]}" for s in items)

def actions_cell(sprint_key, stream_key, stories):
    items = [s for s in stories if s.get('sprint') == sprint_key and s.get('stream') == stream_key]
    if not items: return ''
    groups = {}
    for s in items:
        k = f"{s.get('action','')} ({s.get('persona','')})"
        if k not in groups: groups[k] = {'count': 0, 'pts': 0}
        groups[k]['count'] += 1
        groups[k]['pts'] += float(s.get('pts') or 0)
    return '\n'.join(f"{k}: {v['count']} stories, {v['pts']:.1f}pts" for k, v in groups.items())

def goals_cell(sprint_key, stream_key, stories):
    items = [s for s in stories if s.get('sprint') == sprint_key and s.get('stream') == stream_key]
    if not items: return ''
    groups = {}
    for s in items:
        g = s.get('goal','')
        if g not in groups: groups[g] = {'count': 0, 'pts': 0}
        groups[g]['count'] += 1
        groups[g]['pts'] += float(s.get('pts') or 0)
    return '\n'.join(f"{g}: {v['count']} stories, {v['pts']:.1f}pts" for g, v in groups.items())

make_board_sheet(wb, 'Board - Stories', stories_cell)
make_board_sheet(wb, 'Board - Actions', actions_cell)
make_board_sheet(wb, 'Board - Goals', goals_cell)

# Stream Config sheet
ws5 = wb.create_sheet('Stream Config')
ws5.append(['Stream Number', 'Stream Name', 'Stream Description', 'Stream Color'])
for cell in ws5[1]:
    cell.font = Font(bold=True, color='FFFFFFFF')
    cell.fill = PatternFill('solid', fgColor='FF1E3A5F')
for i, s in enumerate(streams, 1):
    ws5.append([i, s.get('name',''), s.get('description',''), s.get('color','')])
for col_idx, col in enumerate(ws5.columns, 1):
    max_len = max((len(str(c.value or '')) for c in col), default=12)
    ws5.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

wb.save(out_path)
print('ok')
