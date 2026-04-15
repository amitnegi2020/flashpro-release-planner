import json, openpyxl, sys
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

data_path = sys.argv[1]
out_path = sys.argv[2]

with open(data_path) as f:
    data = json.load(f)

stories = data['stories']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Story Map'

# Dynamic headers: fixed system cols + any extra fields
system_keys = ['id', 'headline', 'persona', 'goal', 'action', 'capability', 'workflow', 'pts', 'status']
extra_keys = []
for s in stories:
    for k in s.keys():
        if k not in system_keys and k not in ('sprint', 'stream', 'originalStoryId', 'isSplitChild') and k not in extra_keys:
            extra_keys.append(k)

col_map = [
    ('Story ID', 'id'), ('Headline', 'headline'), ('User Persona', 'persona'),
    ('Goal', 'goal'), ('Action', 'action'), ('Platform Capability', 'capability'),
    ('Workflow', 'workflow'), ('Story Estimate (weeks)', 'pts'), ('Status', 'status'),
] + [(k, k) for k in extra_keys]

headers = [c[0] for c in col_map]
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFFFF')
    cell.fill = PatternFill('solid', fgColor='FF1E3A5F')
    cell.alignment = Alignment(horizontal='center')

for s in stories:
    ws.append([str(s.get(c[1], '') or '') for c in col_map])

for col_idx, col in enumerate(ws.columns, 1):
    max_len = max((len(str(c.value or '')) for c in col), default=10)
    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

wb.save(out_path)
print('ok')
