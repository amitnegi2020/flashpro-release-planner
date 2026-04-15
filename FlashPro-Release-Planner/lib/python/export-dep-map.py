import json, openpyxl, sys
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

data_path = sys.argv[1]
out_path = sys.argv[2]

with open(data_path) as f:
    data = json.load(f)

deps = data.get('customDeps', {})
cp = data.get('criticalPath', [])

wb = openpyxl.Workbook()

ws = wb.active
ws.title = 'Dependencies'
ws.append(['Story ID', 'Depends On'])
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFFFF')
    cell.fill = PatternFill('solid', fgColor='FF1E3A5F')

for sid, dep_list in deps.items():
    for dep in dep_list:
        ws.append([sid, dep])

ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 18

if cp:
    ws2 = wb.create_sheet('Critical Path')
    ws2.append(['Story ID'])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.fill = PatternFill('solid', fgColor='FF7C3AED')
    for sid in cp:
        ws2.append([sid])
    ws2.column_dimensions['A'].width = 18

wb.save(out_path)
print('ok')
